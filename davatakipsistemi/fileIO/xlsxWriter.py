from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl import Workbook



def _create_fill(color: str) -> PatternFill:
    """Create a fill pattern with the given color.

    Args:
        color (str): The color code in hexadecimal format (e.g., 'FFFF00' for yellow).

    Returns:
        PatternFill: A fill pattern object with the specified color.
    """
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def _set_column_width(sheet, column_letters: list, width: int):
    """Set the width of specified columns in the given sheet.

    Args:
        sheet: The Excel sheet where the column width will be set.
        column_letters (list): A list of column letters (e.g., ['A', 'B', 'C']).
        width (int): The width to set for the specified columns.
    """
    for col in column_letters:
        sheet.column_dimensions[col].width = width


def _write_header(sheet, row_index: int, headers: list):
    """Write the header row to the specified sheet with specific styles.

    Args:
        sheet: The Excel sheet where the header will be written.
        row_index (int): The row number where the header will be placed.
        headers (list): A list of header values to write.
    """
    fill_color = _create_fill('66CCFF')  # Blue
    font = Font(size=13, bold=True)
    alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    for col_letter, header in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", headers):
        cell = sheet[f'{col_letter}{row_index}']
        cell.value = header
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment


def _write_row(sheet, row_index: int, data: list, is_header: bool = False):
    """Write a data row to the specified sheet with specific styles.

    Args:
        sheet: The Excel sheet where the data row will be written.
        row_index (int): The row number where the data will be placed.
        data (list): A list of values to write in the row.
        is_header (bool): Indicates if the row being written is a header (default is False).
    """
    fill_color = _create_fill('FFFF00') if not is_header else _create_fill('66CCFF')
    font = Font(size=10)
    alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    for col_letter, value in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", data):
        cell = sheet[f'{col_letter}{row_index}']
        cell.value = value
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment


def write_to_excel(data: list, file_name: str):
    """Write data to an Excel file.

    Args:
        data (list): A list of lists containing the data to be written.
        file_name (str): The name of the file (without the .xlsx extension) to save.
    
    Raises:
        Exception: If the Excel file cannot be saved.
    """
    workbook = Workbook()
    sheet = workbook.create_sheet("Data", 0)

    # Set column widths
    _set_column_width(sheet, "ABCDEFGHIJKLMNOPQRSTUVWXYZ", 30)

    # Write data to the sheet
    for row_index, row_data in enumerate(data, start=1):
        if row_index == 1:
            _write_header(sheet, row_index, row_data)
        else:
            _write_row(sheet, row_index, row_data)

    try:
        workbook.save(f'{file_name}.xlsx')
        print(f"{file_name}.xlsx has been created successfully.")
    except Exception as e:
        print("Please close the Excel file:", e)
